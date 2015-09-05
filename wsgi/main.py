# -*- coding: utf-8 -*-

# Flask functionals Imports
from flask import Flask, render_template, request
from flask import redirect, url_for, session
from functools import wraps
from werkzeug import secure_filename

# DB ORM Imports
from flask_sqlalchemy import SQLAlchemy

# Form, and Form Validators Imports
from flask.ext.wtf import Form
from wtforms import TextField, PasswordField, validators, HiddenField, TextAreaField, BooleanField
from wtforms.validators import Required, EqualTo, Optional, Length, Email, ValidationError
from flask_wtf.file import FileField, FileAllowed, FileRequired
from wtforms import SubmitField

# Login Imports 
from flask.ext.login import LoginManager, login_user, logout_user, current_user, login_required
import os

# Mail Imports
from itsdangerous import URLSafeTimedSerializer
from flask.ext.mail import Mail, Message

#Time Imports
from datetime import datetime

# Global Vars
global filepath

# Excel File Imports
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from openpyxl.drawing import Image

mail = Mail()
application = Flask(__name__)

# upload reports config
application.config['UPLOAD_FOLDER'] = os.environ.get('OPENSHIFT_DATA_DIR') if os.environ.get('OPENSHIFT_DATA_DIR') else 'wsgi/static/patterns'

# cloud and local db
application.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('OPENSHIFT_POSTGRESQL_DB_URL') if os.environ.get('OPENSHIFT_POSTGRESQL_DB_URL') else 'postgres://lomelisan:6c6f6d656c69@localhost:5432/reporta'
application.config['CSRF_ENABLED'] = True
application.config['SECRET_KEY'] = 'esunsecreto'
application.config['SECURITY_PASSWORD_SALT'] = 'esunsecreto'

# login config
login_manager = LoginManager()
login_manager.init_app(application)
login_manager.login_view = '/signin'

# mail config
application.config['MAIL_SERVER'] = 'smtp.gmail.com'
application.config['MAIL_PORT'] = 465
application.config['MAIL_USE_TLS'] = False
application.config['MAIL_USE_SSL'] = True
application.config['MAIL_USERNAME'] = 'reporta.movilnet@gmail.com'
application.config['MAIL_PASSWORD'] = 'esunsecreto'
application.config['MAIL_DEFAULT_SENDER'] = 'reporta-movilnet@gmail.com'

db = SQLAlchemy(application)
mail.init_app(application)

# Global vars
global filepath
global colGen
global countLines
global countApzA1
global heirFilePath
application.config['PATTERNS_FOLDER'] = 'wsgi/static/patterns'



# DB Models
class Users(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(60), unique=True)
    name = db.Column(db.String(40))
    password = db.Column(db.String)
    email = db.Column(db.String(100), unique=True)
    confirmed = db.Column(db.Boolean, nullable=False, default=False)
    role = db.Column(db.String(20))
    

    active = db.Column(db.Boolean)
    
    def __init__(self, username = None, password = None, email = None,
    firstname = None, lastname = None, confirmed = None
    , role = None, active = None):
        self.username = username
        self.email = email
        self.firstname = firstname
        self.lastname = lastname
        self.password = password
        self.confirmed = confirmed
        self.role = role
        self.active = active
        
    def is_authenticated(self):
        return True

    def is_active(self):
        return self.active
    
    def is_anonymous(self):
        return False

    def get_id(self):
        return unicode(self.id)
	


#Token Generator
def generate_confirmation_token(email):
    serializer = URLSafeTimedSerializer(application.config['SECRET_KEY'])
    return serializer.dumps(email, salt=application.config['SECURITY_PASSWORD_SALT'])

#Confirm Token
def confirm_token(token, expiration=3600):
    serializer = URLSafeTimedSerializer(application.config['SECRET_KEY'])
    try:
        email = serializer.loads(
            token,
            salt=application.config['SECURITY_PASSWORD_SALT'],
            max_age=expiration
        )
    except:
        return False
    return email

#Send Email
def send_email(to, subject, template):
    msg = Message(
        subject,
        recipients=[to],
        html=template,
        sender=application.config['MAIL_DEFAULT_SENDER']
    )
    mail.send(msg)
    
#Send Email File
def send_email_file(to, subject, template, path_mail_file, name_file, type_mail_file):
    msg = Message(
        subject,
        recipients=[to],
        html=template,
        sender=application.config['MAIL_DEFAULT_SENDER']
    )
    with application.open_resource(path_mail_file) as fp:
		msg.attach(name_file, type_mail_file, fp.read())
    mail.send(msg)

#Custom decorator
def check_confirmed(func):
    @wraps(func)
    def decorated_function(*args, **kwargs):
        if current_user.confirmed is False:
            return redirect(url_for('unconfirmed'))
        return func(*args, **kwargs)

    return decorated_function
    
#Custom decorator
def check_admin(func2):
    @wraps(func2)
    def decorated_admin_function(*args, **kwargs):
        if current_user.role != 'admin':
            return redirect(url_for('notadmin'))
        return func2(*args, **kwargs)

    return decorated_admin_function

#Custom lowcase Valitador
def lowcase_check(form, field):
	if field.data.isupper() or not field.data.islower():
		raise ValidationError(u'Este campo solo acepta minúsculas')

class SignupForm(Form):
    email = TextField(u'Dirección Email', validators=[
            Required(u'Introduce una dirección email válida'),
            Length(min=6, message=(u'Dirección Email muy corta')),
            Email(message=(u'Dirección Email no válida.')), lowcase_check
            ])
    password = PasswordField('Clave', validators=[
            Required(u'Campo requerido'),
            Length(min=6, message=(u'Introduce una clave mas larga'))           
            ])
    username = TextField('Usuario', validators=[Required(u'Campo Requirido'), lowcase_check])
    
			
    agree = BooleanField(u'Acepto todos los <a href="/static/tos.html">Términos del Servicio</a>', validators=[Required(u'Debes aceptar los términos del servicio')])

	
			
class SigninForm(Form):
    username = TextField('Usuario', validators=[
            Required('Campo Requirido'),
            validators.Length(min=3, message=('Nombre de usuario muy corto')),
            lowcase_check])
    password = PasswordField('Clave', validators=[
            Required('Campo Requirido'),
            validators.Length(min=6, message=('Introduce una clave mas larga'))
            ])
    remember_me = BooleanField(u'Recuérdame', default = False)
    
class UploadForm(Form):
	input_file = FileField('', validators = [
			FileRequired(message = 'No hay archivo para subir!')
			, FileAllowed(['log'], message = 'Solo introduzca archivos .log')
			])
	submit = SubmitField(label = "Subir")
	
class UploadPatternForm(Form):
	input_file = FileField('', validators = [
			FileRequired(message = 'No hay archivo para subir!')
			, FileAllowed(['xlsx', 'txt'], message = 'Solo introduzca archivos .xlsx')
			])
	submit = SubmitField(label = "Subir")
	
	
	

@login_manager.user_loader
def load_user(id):
    return Users.query.get(id)
    
    
@application.route('/')
@application.route('/<username>')
def index(username = None):
    if username is None:
        return render_template('index.html', page_title = 'Inicio', signin_form = SigninForm())
    
    

@application.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        form = SignupForm(request.form)
        if form.validate():
            user = Users()
            form.populate_obj(user)

            user_exist = Users.query.filter_by(username=form.username.data).first()
            email_exist = Users.query.filter_by(email=form.email.data).first()

            if user_exist:
                form.username.errors.append('Usuario ya registrado')

            if email_exist:
                form.email.errors.append('Email ya registrado')
   
							
            if (user_exist or email_exist):
				return render_template('signup.html',
                                       signin_form = SigninForm(),
                                       form = form,
                                       page_title = 'Registro')

            else:
				user.active = True
				user.confirmed = False
				
				db.session.add(user)
				db.session.commit()
				
				token = generate_confirmation_token(user.email)
				confirm_url = url_for('confirm_email', token=token, _external=True)
				html = render_template('activate.html', confirm_url=confirm_url)
				subject = "Por favor confirma tu dirección de correo!"
				send_email(user.email, subject, html)
				
				
				return render_template('signup-success.html',
                                       user = user,
                                       signin_form = SigninForm(),
                                       page_title = 'Registro exitoso!')

        else:
            return render_template('signup.html',
                                   form = form,
                                   signin_form = SigninForm(),
                                   page_title = 'Registro')
    return render_template('signup.html',
                           form = SignupForm(),
                           signin_form = SigninForm(),
                           page_title = 'Registro')

@application.route('/signin', methods=['GET', 'POST'])
def signin():
    if request.method=='POST':
        if current_user is not None and current_user.is_authenticated():
            return redirect(url_for('index'))
    
        form = SigninForm(request.form)
        if form.validate():
            user = Users.query.filter_by(username = form.username.data).first()
            
            if user is None:
                form.username.errors.append('Usuario no registrado')
                return render_template('signinpage.html',  signinpage_form = form,
					page_title = 'Acceso')
            if user.password != form.password.data:
                form.password.errors.append('Clave incorrecta')
                return render_template('signinpage.html',  signinpage_form = form,
					page_title = 'Acceso')          
            
            login_user(user, remember = form.remember_me.data)            

            session['signed'] = True
            session['username']= user.username
            session['email']= user.email
            session['role']= user.role
            
            if session.get('next'):                
                next_page = session.get('next') 
                session.pop('next')
                return redirect(next_page) 
            else:
                return redirect(url_for('index'))
        return render_template('signinpage.html',  signinpage_form = form,
					page_title = 'Acceso')
    else:
        session['next'] = request.args.get('next')
        return render_template('signinpage.html', signinpage_form = SigninForm(),
					page_title = 'Acceso')        


@application.route('/signout')
def signout():
    session.pop('signed')
    session.pop('username')
    session.pop('role')
    session.pop('email')
    logout_user()
    return redirect(url_for('index'))

@application.route('/profile')
@login_required
@check_confirmed
def profile():
    return render_template('profile.html', page_title='Reporta - Perfil')

@application.route('/update_pattern', methods=['GET', 'POST'])
@login_required
@check_confirmed
@check_admin
def update_pattern():
	form = UploadPatternForm()
	if request.method == 'POST' and form.validate_on_submit():
		input_file = request.files['input_file']
		if input_file:
			filename = secure_filename(input_file.filename)
			global filepath 
			filepath = os.path.join(application.config['UPLOAD_FOLDER'], filename)
			input_file.save(filepath)
			return render_template('upload-pattern-success.html', filename=filename, page_title = u'Éxito')
	else:
		return render_template('upload-pattern.html',  uploadpattern_form = form,  page_title = 'Subida')
    
    
@application.route('/adminpanel')
@login_required
@check_confirmed
@check_admin
def adminpanel():
    return render_template('adminpanel.html', page_title='Administrador')
    
@application.route('/confirm/<token>')
@login_required
def confirm_email(token):
    if current_user.confirmed:
		return render_template('already-confirm.html', page_title = 'Cuenta ya Confirmada!')
    else:
		email = confirm_token(token)
		user = Users.query.filter_by(email = current_user.email).first()
		if user.email == email:
			user.confirmed = True
			db.session.add(user)
			db.session.commit()
			return render_template('confirm-success.html', page_title = 'Confirmacion exitosa!')
    return render_template('invalid-confirm.html', page_title = 'Error')

@application.route('/unconfirmed')
@login_required
def unconfirmed():
    if current_user.confirmed:
        return redirect(url_for('index'))
    return render_template('unconfirmed.html')

@application.route('/notadmin')
@login_required
@check_confirmed
def notadmin():
    if current_user.role == 'admin':
        return redirect(url_for('index'))
    return render_template('notadmin.html')
    
@application.route('/upload', methods=['GET', 'POST'])
@login_required
@check_confirmed
def upload():
	form = UploadForm()
	if request.method == 'POST' and form.validate_on_submit():
		input_file = request.files['input_file']
		if input_file:
			filename = secure_filename(input_file.filename)
			global filepath 
			filepath = os.path.join(application.config['UPLOAD_FOLDER'], filename)
			input_file.save(filepath)
			return render_template('upload-success.html', filename=filename, page_title = u'Éxito')
	else:
		return render_template('upload.html', uploadfile_form = form,  page_title = 'Subida')

@application.route('/reading')
@login_required
@check_confirmed
def reading():
	global filepath
	global colGen
	global countApzA1
	global countLines
	global ok_caclp
	global ok_dpwsp
	global ok_plldp
	global ok_exrpp
	global ok_exemp
	global ok_ntstp
	global ok_syrip
	global ok_apamp
	global ok_m3rsp
	global ok_m3asp
	global ok_ihalp
	global ok_ihstp
	global ok_strsp
	global ok_blorp 
	global ok_blurp
	global ok_faiap
	global ok_mgsvp
	global ok_mgarp
	global ok_stbsp
	global ok_sybfp
	
	global ok_prcstate
	global ok_hostname
	global ok_alist
	global ok_cluster_res
	global ok_cluster_group
	global ok_cdhls
	global ok_cdhver
	global ok_afpls
	global ok_aloglist
	
	
	datafile = file(filepath)
	colGen = []
	
	
	
	countLines = 0
	
	allip = False
	countApzA1 = 0
	
	caclp = False
	count_caclp = 0
	mss_date = 0
	mss_time = 0
	day = ' '
	month = ' '
	year = ' '
	hour = ' '
	mins = ' '
	sec = ' '
	y = ' ' 
	
	dpwsp = False
	count_dpwsp = 0
	ok_dpwsp = False
	
	plldp = False
	count_plldp = 0
	col_plldp = []
	col_plldp2 = []
	ok_plldp = False
	
	exrpp= False
	count_exrpp = 0
	ok_exrpp = True
	
	exemp = False
	count_exemp = 0
	ok_exemp = True
	
	ntstp = False
	count_ntstp = 0
	ok_ntstp = False
	
	syrip = False
	count_syrip = 0
	ok_syrip = False
	
	apamp = False
	count_apamp = 0
	count_pas = 0
	count_act = 0
	ctrl_apamp = False
	ok_apamp = False
	
	m3rsp = False
	count_m3rsp = 0
	ok_m3rsp = False
	
	m3asp = False
	count_m3asp = 0
	ok_m3asp = False
	
	ihalp = False
	ok_ihalp = False
	count_ihalp_rip = 0
	count_ihalp_sas = 0
	ctr_ihalp_rip = False
	ctr_ihalp_sas = False
	
	ihstp = False
	count_ihstp = 0
	ok_ihstp = False
	
	strsp = False
	count_strsp = 0
	col_strsp = []
	col_strp2 = []
	b = False
	w = 0
	j = 0
	ok_strsp = False
	
	blorp = False
	count_blorp = 0
	ok_blorp = False
	
	blurp = False
	count_blurp = 0
	ok_blurp = False
	
	faiap = False
	ok_faiap = False
	count_faiap = 0
	
	mgsvp = False
	count_mgsvp = 0
	ok_mgsvp = False
	
	mgarp = False
	ok_mgarp = False
	count_mgarp = 0
	count_mgarp_aux = 0
	col_mgarp = []
	
	stbsp = False
	count_stbsp = 0
	ok_stbsp = False
	
	sybfp = False
	count_ssd = 0
	relfsw0_date = ' ' 
	relfsw0_time = ' '
	ok_sybfp = False
	
	
	
	#----------------------------
	
	prcstate = False
	count_prcstate = 0
	ok_prcstate = False
	
	hostname = False
	count_hostname = 0
	ok_hostname = False
	
	alist = False
	ok_alist = False
	count_alist = 0

	cluster_res = False
	count_cluster_res = 0
	ok_cluster_res = False
	
	cluster_group = False
	count_cluster_group = 0
	
	cdhls = False
	count_cdhls = 0
	ok_cdhls = False
	
	cdhver = False
	count_cdhver = 0
	ok_cdhver = False
	
	afpls = False
	count_afpls_ctr = 0
	count_afpls = 0
	col_afpls = [] 
	col_afpls2 = []
	ok_afpls = False
	
		
	aloglist = False
	count_aloglist = 0
	ok_aloglist = False
	
	#----------
	
	
	
	for line in datafile:
		countLines += 1
		colGen.append(line)
		
		if len(line) == 10:
			continue
		
		if "END" in line:
			allip = False
			caclp = False
			dpwsp = False
			plldp = False
			exrpp = False
			exemp = False
			ntstp = False
			syrip = False
			apamp = False
			m3rsp = False
			m3asp = False
			ihalp = False
			ihstp = False
			strsp = False
			blorp = False
			blurp = False
			faiap = False
			mgsvp = False
			mgarp = False
			stbsp = False
			sybfp = False
		
		#if 	'<allip;' in line:
			#allip = True
		#if allip == True and 'A1/APZ' in line:
			#countApzA1 += 1
		
		
		# Proceso alarma Allip
		if 'A1/APZ' in line:
			countApzA1 += 1
		#----------	 
		
	
		# Proceso alarma Caclp
		if caclp == True:
			count_caclp +=1
			if count_caclp == 5:
				w = line.split()
				mss_date = int(w[1])
				mss_time = int(w[2])
				i = datetime.now().timetuple()
				day = str(i[2])
				month = str(i[1])
				#Agrega 0 a mes o día menor a 10
				if i[2] < 10:
					day = '0' + str(i[2])
	
				if i[1] <10:
					month = '0' + str(i[1])
				#Deja solo ultimos 2 chars al año
				y = str(i[0])
				year = y[2] +y[3]
		
				#Guarda Fecha estilo 150826
				date = int(year + month + day)
				t = date
				hour = str(i[3])
				mins = str(i[4])
				sec = str(i[5])
				#Agrega 0 a hora, min y seg si es menor a 10
				if i[3] < 10:
					hour = '0' + str(i[3])
				if i[4] < 10:
					mins =  '0' + str(i[4])
				if i[5] < 10:
					sec = '0' + str(i[5])
			
				#Guarda Hora estilo 150826
				time = int(hour + mins + sec)
					
				if date == mss_date:
					if time - mss_time > 5 or mss_time - time > 5:
						ok_caclp = False
						caclp = False
					else:
						ok_caclp = True
				else:
					ok_caclp = False
					caclp = False
			
		
		if '<CACLP;' in line:
			caclp = True
			count_caclp +=1
		#----------	
			
			
			
		# Proceso alarma DPWSP	
		if dpwsp == True:
			count_dpwsp += 1
			if 'CPSTATE' in line and count_dpwsp == 2:
				ok_dpwsp = True
			else: 
				ok_dpwsp = False
				
			if 'MAU  SB SBSTATE' in line and count_dpwsp == 3:
				ok_dpwsp = True
			else: 
				ok_dpwsp = False
				
			if 'NRM  B  WO' in line and count_dpwsp == 4:
				ok_dpwsp = True
			else: 
				ok_dpwsp = False
			
		if '<DPWSP;' in line:
			dpwsp = True
			count_dpwsp += 1
		#----------	
		
		# Proceso alarma PLLDP
		if plldp == True:
			count_plldp += 1
			if count_plldp >= 4:
				for i in line:
					if i == ' ':
						if b == True:
							w = "".join(col_plldp2)
							col_plldp.append(w)
							j = 0
							col = []
							b = False
							continue
						else:
							continue
					col_plldp2.append(i)
					b = True
					j += 1
				if col_plldp[2] <= "60":
					ok_plldp = True
				else:
					ok_plldp = False
					plldp = False
				
			
		if '<PLLDP;' in line:
			plldp = True
			count_plldp += 1
		#----------	
		
		# Proceso alarma EXRP
		if exrpp == True:
			count_exrpp += 1
			if count_exrpp >= 4:
				if 'WO' in line:
					ok_exrpp = True
				else :
					exrpp = False
					ok_exrpp = False
					
		
		if '<EXRPP:RP=ALL;' in line:
			exrpp = True
			count_exrpp += 1
		#----------	 
			
		# Proceso alarma EXEMP
		if 	exemp == True:
			count_exemp += 1
			if count_exrpp >= 4:
				if 'WO' in line:
					ok_exemp = True
				else:
					exemp = False
					ok_exemp = False
				
		if '<EXEMP:rp=all,em=ALL;' in line:
			exemp = True
			count_exemp += 1
		#----------
		
			
		# Proceso alarma NTSP
		if 	ntstp == True:
			count_ntstp += 1
			if count_ntstp >= 4:
				if 'WO' in line:
					ok_ntstp = True
				else:
					ntstp = False
					ok_ntstp = False
		
		if '<NTSTP:SNT=ALL;' in line:
			ntstp = True
			count_ntstp += 1
		#----------
		
		# Proceso alarma SYRIP
		if syrip == True:
			count_syrip += 1
			
			if count_syrip >= 7:
				ok_syrip = False
				syrip = False
		
		if '<SYRIP:SURVEY;' in line:
			syrip = True
			count_syrip += 1
		#----------
		
		# Proceso alarma APAMP
		if apamp == True:
			count_apamp += 1
			if count_apamp >= 3:
				if 'PASSIVE' in line:
					count_pas += 1
				if 'ACTIVE' in line:
					count_act += 1
			else:
				apamp = False
					
				
				
		if 'DIRECTORY ADDRESS DATA' in line:
			apamp = True
			ctrl_apamp = True
			count_apamp == 1
		
		if apamp == False and ctrl_apamp == True:
			if count_pas == count_act:
				ok_apamp = True
			else:
				 ok_apamp = False
		#----------
		
		# Proceso alarma M3RSP
		if m3rsp == True:
			count_m3rsp += 1
			if count_m3rsp >= 3:
				w = line.split()
				for i in w:
					if i == 'RST':
						ok_m3rsp = True
						continue
					if i == 'AVA':
						continue
					if i == 'EN-ACT-AVA':
						continue
					else:
						ok_m3rsp = False
						m3rsp = False
					
				
		
		if '<M3RSP:DEST=ALL;' in line:
			m3rsp = True
			count_m3rsp += 1
		#----------
		
		
		# Proceso alarma M3ASP
		if m3asp == True:
			count_m3asp += 1
			if count_m3asp >= 4:
				w = line.split()
				for i in w:
					if i == 'ACT':
						ok_m3asp = True
					else:
						ok_m3asp = False
						m3asp = False
					
		if '<M3ASP;' in line:
			m3asp = True
			count_m3asp += 1
		#----------
		
		
		# Proceso alarma IHALP
		if ihalp == True:
			if ctr_ihalp_sas == True:
				count_ihalp_sas += 1
			
			if 'SASTATE' in line:
				ctr_ihalp_sas = True
				count_ihalp_sas += 1
			
			if count_ihalp_sas == 2:
				ctr_ihalp_sas = False
				count_ihalp_sas = 0
				if 'ASSOCESTABL' in line:
					ok_ihalp = True
				else:
					ok_ihalp = False
					ihalp = False
			
			if ctr_ihalp_rip == True:
				count_ihalp_rip += 1
			
			if 'RIPSTATE' in line:
				ctr_ihalp_rip = True
				count_ihalp_rip += 1
			
			if count_ihalp_rip >=2 and count_ihalp_rip <=3:
				if count_ihalp_rip ==3:
					ctr_ihalp_rip = False
					count_ihalp_rip = 0
					
				w = line.split()
				if w[2] == 'ACTIVE':
					ok_ihalp = True
				else:
					ok_ihalp = False
					ihalp = False
						
					
					
		if '<IHALP:EPID=ALL;' in line:
			ihalp = True
			ok_ihalp = True
		#----------
		
		
		# Proceso alarma IHSTP
		if ihstp == True:
			count_ihstp += 1
			if count_ihstp >= 4:
				if 'BUSY' in line:
					ok_ihstp = True
				else:
					ok_ihstp = False
					ihstp = False
					
		if '<IHSTP:IPPORT=ALL;' in line:
			ihstp = True
			count_ihstp += 1
		#----------
		
		
		# Proceso alarma STRSP
		
		if strsp == True:
			count_strsp += 1
			if count_strsp >= 4:
				w =line.split()
				if w[5] == '0':
					ok_strsp = True
				else:
					ok_strsp = False
					strsp = False		
		
		if '<STRSP:R=ALL;' in line:
			strsp = True
			count_strsp += 1
		#----------
		
		
		
		# Proceso alarma BLORP
		if blorp == True:
			count_blorp += 1
			if count_blorp >= 4:
				if 'NONE' in line:
					ok_blorp = True
					blorp = False
				else:
					ok_blorp = False 
					blorp = False
			
		if '<BLORP;' in line:
			blorp = True
			count_blorp += 1
		#----------
		
		# Proceso alarma BLURP
		if blurp == True:
			count_blurp += 1
			if count_blurp >= 4:
				w =line.split()
				if w[7] == 'YES':
					ok_blrup = True
				else:
					ok_blrup = False
					blurp = False
		
		if '<BLURP:R=ALL;' in line:
			blurp = True
			count_blurp += 1
		#----------
		
		
		# Proceso alarma faiap
		if faiap == True:
			count_faiap += 1
			if count_faiap  > 10:
				ok_faiap = False
				faiap = False
			else:
				ok_faiap = True
		
		if '<FAIAP:R=ALL;' in line:
			faiap = True
			count_faiap += 1
		#----------
		
		
		
		# Proceso alarma mgsvp
		if mgsvp == True:
			count_mgsvp += 1
			if count_mgsvp > 10:
				ok_mgsvp = True
				mgsvp = False
			else:
				ok_mgsvp = False
		
		if '<MGSVP;' in line:
			mgsvp = True
			count_mgsvp += 1	 
		#----------
		
		
		
		# Proceso alarma mgarp
		if mgarp == True:
			count_mgarp += 1
			if count_mgarp%2 == 0 and count_mgarp > 2:
				w = line.split()
				col_mgarp.append(w[1])
			if count_mgarp == 10:
				for i in col_mgarp:
					for j in col_mgarp:
						if i == j:
							count_mgarp_aux += 1
					if count_mgarp_aux >= 2:
						ok_mgarp = False
						mgarp = False
						break
					else:
						ok_mgarp = True
					count_mgarp_aux = 0
					
				
		
		if '<MGARP:NLOG=10;' in line:
			mgarp = True
			count_mgarp += 1
		#----------
		
		
		# Proceso alarma stbsp
		if stbsp == True:
			count_stbsp += 1
			if count_stbsp == 4:
				if 'NONE' in line:
					ok_stbsp = True
				else :
					ok_stbsp = False
					stbsp = False
			
		if '<STBSP:DETY=ALL;' in line:
			stbsp = True
			count_stbsp += 1
		#----------
		
		
		# Proceso alarma sybfp
		if sybfp == True:
			if 'SDD' in line:
				count_ssd += 1
				w = line.split()
				i = w[2]
				j = w[3]
				if count_ssd == 1:
					relfsw0_date = int(i)
					relfsw0_time = int(j)
					
				if count_ssd >= 2:
					if int(i) < relfsw0_date:
						ok_sybfp = True
					else: 
						ok_sybfp = False
					if int(i) == relfsw0_date and int(j) < relfsw0_time:
						ok_sybfp = True
					else: 
						ok_sybfp = False
						
						
		if '<SYBFP:FILE;' in line:
			sybfp = True
		#---------------------------------------------------------------
		
		if '>mml'  in line:
			aloglist = False
		
		
		# Proceso alarma aloglist
		if aloglist == True:
			if count_aloglist == 3:
				w = line.split()
				if w[1] == 'ACTIVE':
					ok_aloglist = True
				else:
					ok_aloglist = False
					aloglist = False
					 
		if '>aloglist' in line:
			afpls = False
			aloglist = True
		#----------
		
		
		# Proceso alarma afpls
		if afpls == True and count_afpls_ctr == 1:
			count_afpls += 1
			if count_afpls >= 4:
				w = line.split()
				col_afpls.append(w[1])
			if count_afpls == 15:
				afpls = False
				count_afpls = 0
		
		if afpls == True and count_afpls_ctr == 2:
			count_afpls += 1
			if count_afpls >= 4:
				w = line.split()
				col_afpls2.append(w[1])
			if count_afpls == 15:
				afpls = False
				count_afpls = 0
						
		if '>afpls' in line:
			count_afpls_ctr += 1
			cdhver = False
			afpls = True
			count_afpls += 1
		
		#----------
		
		
		
		
		# Proceso alarma cdhver
		if cdhver == True:
			count_cdhver += 1
			if count_cdhver == 2:
				if 'DESTINATION' and 'STATUS' in line:
					ok_cdhver = True
				else:
					ok_cdhver = False
			if count_cdhver == 3:
				if 'RTRDEST' and 'OK' in line:
					ok_cdhver = True
				else:
					ok_cdhver = False
			
		if '>cdhver' in line:
			cdhls = False
			cdhver = True
			count_cdhver += 1
		#----------
		
		
		
		# Proceso alarma cdhls
		if cdhls == True:
			count_cdhls += 1
			if count_cdhls == 5:
				if 'RTRDEST' and 'FTPV2' in line:
					ok_cdhls = True
				else:
					ok_cdhls = False
					cdhls = False
					continue
					
			if count_cdhls == 6:
				if 'STSDEST1' and 'FTPV2' in line:
					ok_cdhls = True
					cdhls = False
					continue
					
				
		if '>cdhls' in line:
			cluster_group = False
			cdhls = True
			count_cdhls += 1
		#----------
			
		
		# Proceso alarma cluster group
		if cluster_group == True:
			count_cluster_group += 1
		
			if count_cluster_group >= 5:
				if 'Online' in line:
					ok_cluster_group = True
				else:
					ok_cluster_group = True
					cluster_group = False
		
		
		if '>cluster group' in line:
			cluster_res = False
			cluster_group = True
			count_cluster_group += 1
		#----------	
		
		# Proceso alarma cluster res
		if cluster_res == True:
			count_cluster_res += 1
		
			if count_cluster_res >= 5:
				if 'Online' in line:
					ok_cluster_res = True
				else:
					ok_cluster_res = False
					cluster_res = False
		
		if '>cluster res' in line:
			alist = False
			cluster_res = True
			count_cluster_res += 1
		#----------	
		
		
		# Proceso alarma alist
		if alist == True:
			count_alist += 1
			if count_alist == 2:
				if 'Alarm' in line:
					ok_alist = False
				else:
					ok_alist = True
					
		if '>alist' in line:
			hostname = False
			alist = True
			count_alist += 1
		#----------	
		
		# Proceso alarma hostname
		if hostname == True:
			count_hostname += 1
			if count_hostname == 2:
				if 'MSSVA3APG40B' or 'MSSVA3APG40A' in line:
					ok_hostname = True
				else:
					hostname = False
					ok_hostname = False
				
		if '>hostname' in line:
			prcstate = False
			hostname = True
			count_hostname += 1
		#----------	
		
		# Proceso alarma pcrstate
		if prcstate == True:
			count_prcstate += 1
			if count_prcstate == 2:
				if 'active' in line:
					ok_prcstate = True
				else:
					prcstate = False
					ok_prcstate = False
				
		if '>prcstate' in line:
			prcstate = True
			count_prcstate += 1
		#----------	
		
		
		
		
	os.remove(filepath)	
	
	return render_template('reading-results.html', countLines=countLines,
	page_title = 'Lectura exitosa', t = t)


@application.route('/processing')
@login_required
@check_confirmed
def processing():
	global colGen
	global countApzA1
	global heirFilePath
	global countLines
	global ok_caclp
	global ok_dpwsp
	global ok_plldp
	global ok_exrpp
	global ok_exemp
	global ok_ntstp
	global ok_syrip
	global ok_apamp
	global ok_m3rsp
	global ok_m3asp
	global ok_ihalp
	global ok_ihstp
	global ok_strsp
	global ok_blorp 
	global ok_blurp
	global ok_faiap
	global ok_mgsvp
	global ok_mgarp
	global ok_stbsp
	global ok_sybfp
	
	global ok_prcstate
	global ok_hostname
	global ok_alist
	global ok_cluster_res
	global ok_cluster_group
	global ok_cdhls
	global ok_cdhver
	global ok_afpls
	global ok_aloglist
	patternFilePath = os.path.join(application.config['UPLOAD_FOLDER'], "ModeloD.xlsx")
	heirFilePath = os.path.join(application.config['UPLOAD_FOLDER'], "reporta/reporta.xlsx")
	i = 0
	j = 0
	wb = load_workbook(patternFilePath)
	ws = wb.get_sheet_by_name("logo")
	ws2 = wb.get_sheet_by_name("MSS")
	rows = countLines
	x = "reporta.xlsx#logo!A1"
	
	ok = Font(name='Arial', size= 8,color=colors.BLACK, bold=True)
	not_ok = Font(name='Arial', size= 8,color=colors.RED, bold=True)	
	for i in range(rows):
		ws.cell(row=i+1, column=1).value = colGen[i]
	j=5
	i = 0
	rows = 57
	for i in range(rows):
		if i  == 17:
			ws2.cell(row=i+1, column=j+1).hyperlink = (x)
			ws2.cell(row=i+1, column=j+1).value = '+Inf'
			if countApzA1 >= 1:
				ws2.cell(row=i+1, column=j).value = "NOT OK"
				ws2.cell(row=i+1, column=j).font = not_ok	
			else:
				ws2.cell(row=i+1, column=j).value = "OK"
				ws2.cell(row=i+1, column=j).font = ok	
			
		if i >= 23 and i <= 42:
			ws2.cell(row=i+1, column=j+1).hyperlink = (x)
			ws2.cell(row=i+1, column=j+1).value = '+Inf'
			if i == 23:
				if ok_caclp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 24:
				if ok_dpwsp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 25:
				if ok_plldp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 26:
				if ok_exrpp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 27:
				if ok_exemp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 28:
				if ok_ntstp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 29:
				if ok_syrip == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok	
			if i == 30:
				if ok_apamp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 31:
				if ok_m3rsp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 32:
				if ok_m3asp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 33:
				if ok_ihalp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 34:
				if ok_ihstp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 35:
				if ok_strsp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 36:
				if ok_blorp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 37:
				if ok_blurp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 38:
				if ok_faiap == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok	
			if i == 39:
				if ok_mgsvp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok	
			if i == 40:
				if ok_mgarp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 41:
				if ok_stbsp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 42:
				if ok_sybfp == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
					
					
		if i == 57:
			break
		if i >= 48:
			ws2.cell(row=i+1, column=j+1).hyperlink = (x)
			ws2.cell(row=i+1, column=j+1).value = '+Inf'
			if i == 48:
				if ok_prcstate == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok	
			if i == 49:
				if ok_hostname == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok	
			if i == 50:
				if ok_alist == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 51:
				if ok_cluster_res == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 52:
				if ok_cluster_group == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 53:
				if ok_cdhls == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok	
			if i == 54:
				if ok_cdhver == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 55:
				if ok_afpls == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
			if i == 56:
				if ok_aloglist == True:
					ws2.cell(row=i+1, column=j).value = "OK"
					ws2.cell(row=i+1, column=j).font = ok
				else:
					ws2.cell(row=i+1, column=j).value = "NOT OK"
					ws2.cell(row=i+1, column=j).font = not_ok
				
			
	
	wb.save(heirFilePath)
	
	return render_template('processing-results.html', page_title = 'Proceso exitoso')
	
@application.route('/sending')
@login_required
@check_confirmed
def sending():		
	global heirFilePath
	#File sender
	if os.environ.get('OPENSHIFT_DATA_DIR'):
		path_mail_file = heirFilePath
	else:
		path_mail_file = "static/patterns/reporta/reporta.xlsx"
	type_mail_file = "excel/xlsx"
	name_file = "reporta.xlsx"
	html = render_template('report.html')
	subject = "Has recibido un Reporte!"
	send_email_file(current_user.email, subject, html, path_mail_file, name_file, type_mail_file)
	
	os.remove(heirFilePath)
	
	return render_template('sending-results.html', 
	page_title = 'Reporte enviado' )



def dbinit():
	db.drop_all()
	db.create_all()
	admin = Users(username='lomelisan', firstname='Carlos', 
		lastname='Lomeli', password='esunsecreto', 
		email='lomelisan@hotmail.com', confirmed = True, 
		role='admin', active = True)
	db.session.add(admin)
	db.session.commit()




if __name__ == '__main__':
	dbinit()
	application.run(debug=True, host="0.0.0.0", port=8888)





